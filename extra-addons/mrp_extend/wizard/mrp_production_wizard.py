from odoo import fields, models, _
from odoo.exceptions import UserError, MissingError
from odoo.modules import get_module_path
from odoo.http import request
import base64
import io
from openpyxl import load_workbook
import os

class MrpProductionWiz(models.TransientModel):
    _name = 'mrp.product.wiz'
    _description = 'Mrp Production Wizard'

    file_import = fields.Binary(string='File import', help='File import to tab component. Only support Excel file', tracking=True, attachment=False)
    file_name = fields.Char(string='File name', tracking=True)

    def action_import(self):
        active_id = self.env.context.get('active_id')
        if not active_id:
            raise MissingError(_("Manufacturing order records are not found."))
        mo = self.env['mrp.production'].browse(active_id)

        if not self.file_import:
            raise FileNotFoundError(_("No file. Please import file."))
        if not self.file_name or not self.file_name.lower().endswith(('.xls', '.xlsx')):
            raise UserError(_("File import isn't file excel. Please import file excel."))
        
        file_data = io.BytesIO(base64.b64decode(self.file_import)) # Mã hóa file và đưa vào bộ nhớ đệm để xử lý
        try:
            workbook = load_workbook(file_data, data_only=True) # Mở file
        except:
            raise UnicodeDecodeError(_("Error can not read the file."))
        
        sheet = workbook.active
        lines = []

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            product_name, quantity = row
            if not product_name:
                continue
            product = self.env['product.product'].search([
                ('name', '=', product_name),
            ], limit=1)
            if not product:
                raise UserError(_(f"Name of the product: '{product_name}' (row:{idx}) in the file is not in the product list"))
            # Tạo mới bản ghi
            lines.append((0, 0, {
                'product_id': product.id,
                'name': product_name,
                'product_uom_qty': quantity,
            }))
        mo.move_raw_ids = [(5, 0, 0)] + lines # Xóa toàn bộ dữ liệu cũ và ghi đè dữ liệu mới

    def action_download_template(self):
        return {
            'type': 'ir.actions.act_url',
            'url': '/mrp_production/template_import_lines',
            'target': 'self',
        }
